import io
import re
import tempfile
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pdfplumber
import streamlit as st

# ── core logic (stesso algoritmo di bilancio_to_excel.py) ────────────────────

def get_unique_chars(page):
    seen, result = set(), []
    for c in page.chars:
        key = (round(c["x0"], 1), round(c["top"], 1), c["text"])
        if key not in seen:
            seen.add(key)
            result.append(c)
    return result


def chars_to_rows(chars, y_tol=2):
    rows = defaultdict(list)
    for c in chars:
        y = round(c["top"] / y_tol) * y_tol
        rows[y].append(c)
    return {y: sorted(v, key=lambda c: c["x0"]) for y, v in rows.items()}


def parse_it_number(s):
    s = s.strip()
    if not s:
        return None
    negative = s.endswith("-")
    s = s.rstrip("-").strip().replace(".", "").replace(",", ".")
    try:
        val = float(s)
        return -val if negative else val
    except ValueError:
        return None


def detect_saldo_columns(pdf):
    for page in pdf.pages[:3]:
        chars = get_unique_chars(page)
        rows = chars_to_rows(chars)
        for _, line_chars in sorted(rows.items()):
            ns = [(c["text"], c["x0"]) for c in line_chars if c["text"].strip()]
            text_only = "".join(t for t, _ in ns)
            if "Codice" in text_only and "Saldo" in text_only and "Progressivo" in text_only:
                matches = [m.start() for m in re.finditer("Saldo", text_only)]
                if len(matches) >= 3:
                    return ns[matches[0]][1], ns[matches[-2]][1], ns[matches[-1]][1]
                elif len(matches) == 2:
                    return ns[matches[0]][1], ns[matches[-2]][1], ns[matches[-1]][1]
    return None, None, None


def convert(pdf_bytes: bytes) -> bytes:
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = Path(tmp.name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bilancio"

    ws.append(["Codice", "Descrizione", "Saldo"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    with pdfplumber.open(tmp_path) as pdf:
        num_start_x, saldo_dare_x, saldo_avere_x = detect_saldo_columns(pdf)

        if not saldo_dare_x:
            w = pdf.pages[0].width
            num_start_x   = w * 0.490
            saldo_dare_x  = w * 0.790
            saldo_avere_x = w * 0.875

        for page in pdf.pages:
            chars = get_unique_chars(page)
            rows  = chars_to_rows(chars)

            for _, line_chars in sorted(rows.items()):
                raw = "".join(c["text"] for c in line_chars).strip()
                if not raw or set(raw.replace(" ", "")) <= {"-", "="}:
                    continue

                ns_left = [(c["text"], c["x0"]) for c in line_chars
                           if c["x0"] < num_start_x and c["text"].strip()]
                if not ns_left:
                    continue
                left_text = "".join(t for t, _ in ns_left)

                m = re.match(r"^(\d{2,3}(?:\.\d{2}(?:\.\d{4,6})?)?)", left_text)
                if not m:
                    continue
                code = m.group(1)

                code_end_x = ns_left[len(code) - 1][1] + 6
                desc_chars = [c for c in line_chars
                              if code_end_x <= c["x0"] < num_start_x]
                desc = re.sub(r"^[\-\s*]+", "", "".join(c["text"] for c in desc_chars)).strip()

                dare_str = "".join(
                    c["text"] for c in line_chars
                    if saldo_dare_x <= c["x0"] < saldo_avere_x
                ).strip()
                avere_str = "".join(
                    c["text"] for c in line_chars
                    if c["x0"] >= saldo_avere_x
                ).strip()

                dare_val  = parse_it_number(dare_str)
                avere_val = parse_it_number(avere_str)

                if dare_val:
                    saldo = dare_val
                elif avere_val:
                    saldo = -abs(avere_val)
                else:
                    saldo = 0

                ws.append([code, desc, saldo])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    tmp_path.unlink(missing_ok=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UI ───────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="PDF → Excel", page_icon="📊", layout="centered")

st.title("📊 Bilancio di Verifica — PDF → Excel")
st.write("Carica un PDF con tabelle e scarica il file Excel con Codice, Descrizione e Saldo.")

uploaded = st.file_uploader("Scegli un file PDF", type="pdf")

if uploaded:
    st.info(f"File caricato: **{uploaded.name}**")
    if st.button("Converti in Excel", type="primary"):
        with st.spinner("Elaborazione in corso..."):
            try:
                excel_bytes = convert(uploaded.read())
                out_name = Path(uploaded.name).stem + ".xlsx"
                st.success("Conversione completata!")
                st.download_button(
                    label="⬇️ Scarica Excel",
                    data=excel_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Errore durante la conversione: {e}")
